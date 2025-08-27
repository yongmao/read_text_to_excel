# Press the green button in the gutter to run the script.
import openpyxl
import line_to_cell
import datetime
# from prompt_toolkit.shortcuts import yes_no_dialog


def convert():
    # source
    workbook_src = openpyxl.load_workbook('./xl/wxbg.xlsx')
    sheet_src = workbook_src["微信出车报告"]

    rows_src = sheet_src.max_row + 1

    # 列标识, 以1开始
    columns_src_id = 1         # id
    columns_src_date = 2        # 日期
    columns_src_sender = 3      # 发送人
    columns_src_data = 4        # 报告
    columns_src_status = 5      # 导入状态

    # destination
    workbook_dest = openpyxl.load_workbook('./xl/ccjl.xlsx')
    sheet_dest = workbook_dest["details"]
    row_dest_start = sheet_dest.max_row

    columns_dest_data = 26
    sum = 0
    row_dest = row_dest_start + 1
    for i_row_src,row_src in enumerate(sheet_src.iter_rows(values_only=True)):
        if i_row_src == 0:
            continue


        # status = row_src[columns_src_status]
        status = sheet_src.cell(i_row_src + 1, columns_src_status).value
        id = sheet_src.cell(i_row_src + 1, columns_src_id).value
        if status == "yes":
            continue
        data = sheet_src.cell(i_row_src + 1, columns_src_data).value

        # id
        id_src = datetime.datetime.now().strftime('%Y%m%d-%H%M%S') + str(datetime.datetime.now().timestamp()).split(".")[1][:3]
        date_src = datetime.datetime.now().strftime('%Y/%m/%d')
        
        print("\n========= " + id_src + " 开始处理 =========")

        # 拆分成一行一行
        cell_lines = data.split('\n')

        # 循环读取每一份报告的每一行
        # print("=============================")
        for cell_line in cell_lines:
            cell = line_to_cell.parse(cell_line)
            # print(cell)
            if len(cell) > 0:
                sheet_dest.cell(row_dest, cell[0]).value = cell[1]
        #  end for cell_line in cell_lines:

        sheet_dest.cell(row_dest, columns_dest_data + 1).value = data
        sheet_dest.cell(row_dest, 1).value = id_src

        row_dest += 1

        sheet_src.cell(i_row_src + 1, columns_src_id).value = id_src
        sheet_src.cell(i_row_src + 1, columns_src_date).value = date_src
        sheet_src.cell(i_row_src + 1, columns_src_status).value = "yes"
        sum += 1
        
        print("========= " + id_src + " 结束处理 =========")
    # save source excel
    # 覆盖原来的文件，特别是wxbg。 但为测试，另存新文件
    workbook_src.save("./xl/wxbg_new.xlsx")
    workbook_dest.save("./xl/ccjl_new.xlsx")
    
    # 应该覆盖原来的文件
    # workbook_src.save("./xl/wxbg.xlsx")
    # workbook_dest.save("./xl/ccjl.xlsx")
    
    print("\n\n")
    print("  导入了 " + str(sum) + " 条记录！")
    
if __name__ == '__main__':
    # print_hi('PyCharm')
    convert()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
